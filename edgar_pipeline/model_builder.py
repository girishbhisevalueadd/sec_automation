"""Excel model builder using openpyxl.

Multi-sheet workbook: Cover, Income, Balance, Cashflow, Debt, Ratios, Data.
Styling: dark-blue headers, banded rows, frozen panes, Indian number format.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Style primitives
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor=config.COLOR_HEADER_FILL)
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color=config.COLOR_HEADER_TEXT)
SUBTOTAL_FILL = PatternFill("solid", fgColor=config.COLOR_SUBTOTAL_FILL)
SUBTOTAL_FONT = Font(name="Calibri", size=11, bold=True)
ALT_ROW_FILL = PatternFill("solid", fgColor=config.COLOR_ALT_ROW_FILL)
THIN_BORDER = Border(top=Side(style="thin"))
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


SUBTOTAL_KEYWORDS = [
    "total", "subtotal", "gross profit", "operating income",
    "net income", "ebitda", "stockholders equity",
]


def _combine_frames(*frames: pd.DataFrame) -> pd.DataFrame:
    """Stack multiple statement DataFrames into one (for combined sheets).
    Preserves row order; aligns period columns.
    """
    non_empty = [f for f in frames if f is not None and not f.empty]
    if not non_empty:
        return pd.DataFrame()
    if len(non_empty) == 1:
        return non_empty[0]
    combined = pd.concat(non_empty, axis=0, join="outer")
    if "label" in combined.columns:
        cols = ["label"] + [c for c in combined.columns if c != "label"]
        combined = combined[cols]
    return combined


def _is_subtotal(label: str) -> bool:
    if not label:
        return False
    low = str(label).lower()
    return any(k in low for k in SUBTOTAL_KEYWORDS)


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER if c > 1 else LEFT
    ws.row_dimensions[row].height = 22


def _autosize(ws: Worksheet, ncols: int, cap: int = 30) -> None:
    for col_idx in range(1, ncols + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is None:
                continue
            max_len = max(max_len, min(len(str(val)) + 2, cap))
        ws.column_dimensions[letter].width = max_len


def _write_footer(ws: Worksheet, start_row: int, ncols: int) -> None:
    if ncols < 1:
        ncols = 1
    ws.cell(row=start_row, column=1, value=config.DATA_SOURCE_FOOTER).font = Font(
        italic=True, size=9, color="666666"
    )
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A{start_row}:{last_col}{start_row}")


def _write_statement_sheet(
    ws: Worksheet, title: str, df: pd.DataFrame, add_yoy: bool = True
) -> None:
    """Render a statement DataFrame on a worksheet with styling.

    Expects `df` with a 'label' column plus period columns (most recent first).
    """
    ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True, color=config.COLOR_HEADER_FILL)

    if df is None or df.empty:
        ws.cell(row=3, column=1, value="No data available.")
        _write_footer(ws, 5, 1)
        return

    cols = list(df.columns)
    # Make sure 'label' is first
    if "label" in cols:
        cols.remove("label")
        cols = ["label"] + cols
    df = df[cols]

    period_cols = [c for c in cols if c != "label"]
    # Optional YoY column based on first two periods
    yoy_col_letter = None
    header_row = 3
    ws.cell(row=header_row, column=1, value="Line Item")
    for i, c in enumerate(period_cols, start=2):
        ws.cell(row=header_row, column=i, value=str(c))
    last_data_col = 1 + len(period_cols)
    if add_yoy and len(period_cols) >= 2:
        last_data_col += 1
        ws.cell(row=header_row, column=last_data_col, value="YoY %")
        yoy_col_letter = get_column_letter(last_data_col)
    _style_header_row(ws, header_row, last_data_col)

    # Data rows
    for r_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        label = str(row.get("label", ""))
        ws.cell(row=r_idx, column=1, value=label).alignment = LEFT
        for c_idx, c in enumerate(period_cols, start=2):
            val = row[c]
            try:
                fval = float(val) if val not in (None, "") and pd.notna(val) else None
            except (TypeError, ValueError):
                fval = None
            cell = ws.cell(row=r_idx, column=c_idx, value=fval)
            cell.number_format = config.INDIAN_NUMBER_FORMAT
            cell.alignment = RIGHT

        if yoy_col_letter and len(period_cols) >= 2:
            cur_letter = get_column_letter(2)
            prev_letter = get_column_letter(3)
            formula = f"=IFERROR(({cur_letter}{r_idx}-{prev_letter}{r_idx})/ABS({prev_letter}{r_idx}),\"\")"
            yoy_cell = ws.cell(row=r_idx, column=last_data_col, value=formula)
            yoy_cell.number_format = "0.00%"
            yoy_cell.alignment = RIGHT

        is_sub = _is_subtotal(label)
        if is_sub:
            for c in range(1, last_data_col + 1):
                cc = ws.cell(row=r_idx, column=c)
                cc.fill = SUBTOTAL_FILL
                cc.font = SUBTOTAL_FONT
                cc.border = THIN_BORDER
        elif r_idx % 2 == 0:
            for c in range(1, last_data_col + 1):
                ws.cell(row=r_idx, column=c).fill = ALT_ROW_FILL

    ws.freeze_panes = "B4"
    _autosize(ws, last_data_col)

    # Footer
    footer_row = ws.max_row + 2
    _write_footer(ws, footer_row, last_data_col)


def _write_cover(ws: Worksheet, ticker: str) -> None:
    ws.cell(row=2, column=2, value=f"{ticker.upper()} - Financial Model").font = Font(
        size=22, bold=True, color="FFFFFF"
    )
    ws.cell(row=2, column=2).fill = HEADER_FILL
    ws.merge_cells("B2:F2")
    ws.cell(row=4, column=2, value="Prepared by:").font = Font(bold=True)
    ws.cell(row=4, column=3, value=getattr(config, "REPORT_AUTHOR", "Vivek Pol")).font = Font(bold=True, size=12)
    ws.cell(row=5, column=2, value="Company:").font = Font(bold=True)
    ws.cell(row=5, column=3, value=getattr(config, "REPORT_COMPANY", "ValueAdd Research And Analytics Solutions LLP"))
    ws.cell(row=6, column=2, value="Report date:").font = Font(bold=True)
    ws.cell(row=6, column=3, value=datetime.utcnow().strftime("%Y-%m-%d"))
    ws.cell(row=7, column=2, value="Data source:").font = Font(bold=True)
    ws.cell(row=7, column=3, value="SEC EDGAR via edgartools")
    ws.cell(row=9, column=2, value="Sheets in this workbook:").font = Font(bold=True)
    contents = [
        "1. Cover - this page",
        "2. Income Statement",
        "3. Balance Sheet",
        "4. Cash Flow Statement",
        "5. Debt Schedule (incl. Maturity Ladder)",
        "6. Segment & Geographic Detail",
        "7. Lease Commitments",
        "8. Stock-Based Compensation",
        "9. Income Tax Detail",
        "10. Key Ratios",
        "11. Data (raw normalized dump for audit)",
    ]
    for i, line in enumerate(contents, start=10):
        ws.cell(row=i, column=2, value=line)
    for col_letter in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[col_letter].width = 22

    # Embed the logo in the top-right of the cover sheet
    logo_path = getattr(config, "LOGO_PATH", None)
    if logo_path and Path(logo_path).exists():
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(str(logo_path))
            # Resize to ~180px wide while keeping aspect ratio
            try:
                ratio = img.height / img.width
                img.width = 180
                img.height = int(180 * ratio)
            except Exception:  # noqa: BLE001
                pass
            ws.add_image(img, "F2")
        except Exception as e:  # noqa: BLE001
            logger.warning("Could not embed logo on Excel cover (%s): %s", logo_path, e)

    _write_footer(ws, 22, 6)


def _write_ratios_sheet(ws: Worksheet, ratios_df: pd.DataFrame) -> None:
    ws.cell(row=1, column=1, value="Key Ratios").font = Font(size=14, bold=True, color=config.COLOR_HEADER_FILL)
    if ratios_df is None or ratios_df.empty:
        ws.cell(row=3, column=1, value="No ratio data computable from available filings.")
        _write_footer(ws, 5, 1)
        return

    df = ratios_df.reset_index()
    cols = list(df.columns)
    # Trend column using ▲▼ comparing first two period columns (period columns are everything but 'Metric')
    period_cols = [c for c in cols if c != "Metric"]
    header_row = 3
    ws.cell(row=header_row, column=1, value="Metric")
    for i, c in enumerate(period_cols, start=2):
        ws.cell(row=header_row, column=i, value=str(c))
    trend_col_idx = 2 + len(period_cols)
    ws.cell(row=header_row, column=trend_col_idx, value="Trend")
    _style_header_row(ws, header_row, trend_col_idx)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        ws.cell(row=r_idx, column=1, value=str(row["Metric"])).alignment = LEFT
        for c_idx, c in enumerate(period_cols, start=2):
            v = row[c]
            try:
                fv = float(v) if pd.notna(v) else None
            except (TypeError, ValueError):
                fv = None
            cell = ws.cell(row=r_idx, column=c_idx, value=fv)
            cell.number_format = "#,##0.00"
            cell.alignment = RIGHT
        # Trend marker
        if len(period_cols) >= 2:
            try:
                a = float(row[period_cols[0]])
                b = float(row[period_cols[1]])
                marker = "▲" if a > b else ("▼" if a < b else "▬")
            except (TypeError, ValueError):
                marker = ""
            tcell = ws.cell(row=r_idx, column=trend_col_idx, value=marker)
            tcell.alignment = CENTER
            tcell.font = Font(bold=True, color="2E7D32" if marker == "▲" else "C62828" if marker == "▼" else "555555")
        if r_idx % 2 == 0:
            for c in range(1, trend_col_idx + 1):
                if ws.cell(row=r_idx, column=c).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=r_idx, column=c).fill = ALT_ROW_FILL

    ws.freeze_panes = "B4"
    _autosize(ws, trend_col_idx)
    _write_footer(ws, ws.max_row + 2, trend_col_idx)


def _write_data_sheet(ws: Worksheet, summary_dict: dict[str, pd.DataFrame]) -> None:
    ws.cell(row=1, column=1, value="Raw Normalized Data Dump").font = Font(size=14, bold=True)
    row_cursor = 3
    for stmt_name, df in summary_dict.items():
        ws.cell(row=row_cursor, column=1, value=stmt_name.upper()).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row_cursor, column=1).fill = HEADER_FILL
        row_cursor += 1
        if df is None or df.empty:
            ws.cell(row=row_cursor, column=1, value="(no data)")
            row_cursor += 2
            continue
        cols = list(df.columns)
        for c_idx, c in enumerate(cols, start=1):
            ws.cell(row=row_cursor, column=c_idx, value=str(c)).font = Font(bold=True)
        row_cursor += 1
        for _, r in df.iterrows():
            for c_idx, c in enumerate(cols, start=1):
                v = r[c]
                if isinstance(v, float):
                    cell = ws.cell(row=row_cursor, column=c_idx, value=v)
                    cell.number_format = "#,##0"
                else:
                    ws.cell(row=row_cursor, column=c_idx, value=str(v) if pd.notna(v) else "")
            row_cursor += 1
        row_cursor += 1
    _autosize(ws, 8)
    _write_footer(ws, row_cursor + 1, 8)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def build_excel_model(
    ticker: str,
    summary_dict: dict[str, pd.DataFrame],
    ratios_df: pd.DataFrame,
    form_type: str = "10-K",
) -> Path:
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    _write_cover(cover, ticker)

    _write_statement_sheet(
        wb.create_sheet("Income Statement"),
        f"{ticker.upper()} Income Statement",
        summary_dict.get("income", pd.DataFrame()),
    )
    _write_statement_sheet(
        wb.create_sheet("Balance Sheet"),
        f"{ticker.upper()} Balance Sheet",
        summary_dict.get("balance", pd.DataFrame()),
    )
    _write_statement_sheet(
        wb.create_sheet("Cash Flow Statement"),
        f"{ticker.upper()} Cash Flow Statement",
        summary_dict.get("cashflow", pd.DataFrame()),
    )
    # Combine debt + debt_maturity for a fuller Debt Schedule sheet
    debt_combined = _combine_frames(
        summary_dict.get("debt", pd.DataFrame()),
        summary_dict.get("debt_maturity", pd.DataFrame()),
    )
    _write_statement_sheet(
        wb.create_sheet("Debt Schedule"),
        f"{ticker.upper()} Debt Schedule (incl. Maturity Ladder)",
        debt_combined,
        add_yoy=False,
    )
    _write_statement_sheet(
        wb.create_sheet("Segment & Geography"),
        f"{ticker.upper()} Segment & Geographic Detail",
        summary_dict.get("segment", pd.DataFrame()),
        add_yoy=False,
    )
    _write_statement_sheet(
        wb.create_sheet("Lease Schedule"),
        f"{ticker.upper()} Lease Commitments",
        summary_dict.get("leases", pd.DataFrame()),
        add_yoy=False,
    )
    _write_statement_sheet(
        wb.create_sheet("Stock-Based Comp"),
        f"{ticker.upper()} Stock-Based Compensation",
        summary_dict.get("sbc", pd.DataFrame()),
        add_yoy=False,
    )
    _write_statement_sheet(
        wb.create_sheet("Tax Detail"),
        f"{ticker.upper()} Income Tax Detail",
        summary_dict.get("tax_detail", pd.DataFrame()),
        add_yoy=False,
    )
    _write_ratios_sheet(wb.create_sheet("Key Ratios"), ratios_df)
    _write_data_sheet(wb.create_sheet("Data"), summary_dict)

    today = datetime.utcnow().strftime("%Y%m%d")
    form_safe = form_type.replace("-", "")
    out_path = config.EXCEL_DIR / f"{ticker.upper()}_{form_safe}_{today}.xlsx"
    wb.save(out_path)
    logger.info("Excel model saved -> %s", out_path)
    return out_path
